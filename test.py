import time
from bs4 import BeautifulSoup
import requests
import openpyxl

wb=openpyxl.load_workbook('1000companydata.xlsx',data_only=True)
ws=wb.active

for i in range(100):
    print('page : ',i)

    req = requests.get('http://m.mk.co.kr/yearbook/index.php?page='+str(i)+'&TM=Y2&MM=T0')
    req.encoding=None

    html=req.text

    soup= BeautifulSoup(html,'html.parser')
    tds=soup.find("tbody")
    companies=tds.find_all('a')
    ranking=tds.find_all('td',{'class':'center'})

    company=[]
    ranks=[]

    for idx,a in enumerate(companies):
        b=a.text
        print(b,i*10+idx+1)

        b=b.replace('(주)','')
        company.append(b)

    for rank in ranking:
        ranks.append(rank.text)

    for j in range(10):
        ws.cell(row=i * 10 + j + 2, column=1).value = company[j]
        ws.cell(row=i * 10 + j + 2, column=2).value = int(ranks[j])

    wb.save('1000companydata.xlsx')
    time.sleep(3)
    print('page'+str(i)+'done.')